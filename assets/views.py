from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import openpyxl
from django.utils import timezone
import datetime
from openpyxl.styles import Font, PatternFill

from accounts.views import roles_required
from assets.factories import AssetFactory
from .models import Asset
from .forms import AssetForm
from requests.models import AssetRequest
from django.db.models import Count
import json
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.db.models import Q


@login_required
def admin_manage_assets(request):
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', 'all')
    condition_filter = request.GET.get('condition', 'all')

    assets = Asset.objects.all().order_by('-created_at')

    if search_query:
        assets = assets.filter(
            Q(asset_category__icontains=search_query) |
            Q(model__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(barcode__icontains=search_query)
        )

    if status_filter != 'all':
        assets = assets.filter(status=status_filter)

    if condition_filter != 'all':
        assets = assets.filter(asset_condition=condition_filter)

    paginator = Paginator(assets, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'condition_filter': condition_filter,
    }
    return render(request, 'assets/admin_manage_assets.html', context)

@login_required
def admin_asset_detail(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    return render(request, 'assets/admin_asset_detail.html', {'asset': asset})

@login_required
def admin_add_asset(request):
    if request.method == 'POST':
        form = AssetForm(request.POST)
        if form.is_valid():
            asset = form.save(commit=False)
            asset.created_by = request.user
            asset.save()
            messages.success(request, 'Asset added successfully!')
            return redirect('assets:admin_manage_assets')  # UPDATED
    else:
        form = AssetForm()
    return render(request, 'assets/admin_asset_form.html', {'form': form, 'title': 'Add Asset'})  

@login_required
def admin_edit_asset(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        form = AssetForm(request.POST, instance=asset)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asset updated successfully!')
            return redirect('assets:admin_manage_assets')  # UPDATED
    else:
        form = AssetForm(instance=asset)
    return render(request, 'assets/admin_asset_form.html', {'form': form, 'title': 'Edit Asset'}) 


@login_required
def admin_delete_asset(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        asset.delete()
        messages.success(request, 'Asset deleted successfully!')
        return redirect('assets:admin_manage_assets')  # UPDATED
    return render(request, 'assets/admin_asset_confirm_delete.html', {'asset': asset})



@login_required
# def admin


@login_required
def staff_manage_assets(request):
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', 'all')
    condition_filter = request.GET.get('condition', 'all')

    assets = Asset.objects.all().order_by('-created_at')

    if search_query:
        assets = assets.filter(
            Q(asset_category__icontains=search_query) |
            Q(model__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(barcode__icontains=search_query)
        )

    if status_filter != 'all':
        assets = assets.filter(status=status_filter)

    if condition_filter != 'all':
        assets = assets.filter(asset_condition=condition_filter)

    paginator = Paginator(assets, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'condition_filter': condition_filter,
    }
    return render(request, 'assets/staff_manage_assets.html', context)


@login_required
def asset_detail(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    return render(request, 'assets/asset_detail.html', {'asset': asset})


@login_required
def add_asset(request):
    if request.method == 'POST':
        form = AssetForm(request.POST)
        if form.is_valid():
            asset = form.save(commit=False)
            asset.created_by = request.user
            asset.save()
            messages.success(request, 'Asset added successfully!')
            return redirect('assets:manage_assets')  # UPDATED
    else:
        form = AssetForm()
    return render(request, 'assets/asset_form.html', {'form': form, 'title': 'Add Asset'})


@login_required
def edit_asset(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        form = AssetForm(request.POST, instance=asset)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asset updated successfully!')
            return redirect('assets:staff_manage_assets') 
    else:
        form = AssetForm(instance=asset)
    return render(request, 'assets/asset_form.html', {'form': form, 'title': 'Edit Asset'}) 


@login_required
def export_report_excel(request, report_type):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    username = request.GET.get('username')  # Only used for request_summary

    # -----------------------------
    # Safe date parsing
    # -----------------------------
    def parse_date(date_str):
        if not date_str:
            return None
        try:
            return datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return None

    start_date = parse_date(start_date)
    end_date = parse_date(end_date)

    # -----------------------------
    # Workbook setup
    # -----------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color="FFCCEEFF", end_color="FFCCEEFF", fill_type="solid")
    header_font = Font(bold=True)

    # ============================================================
    # 🔹 REPORT TYPE: ASSET USAGE
    # ============================================================
    if report_type == 'asset_usage':
        ws.title = "Asset Usage Report"

        headers = ["Asset Category", "Model", "Serial Number",
                   "Total Requests", "Approved Requests"]
        ws.append(headers)

        for col, _ in enumerate(headers, 1):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).fill = header_fill

        for asset in Asset.objects.all():

            requests_qs = asset.assigned_requests.all()   # ✔ FIXED REVERSE FK

            if start_date:
                requests_qs = requests_qs.filter(request_date__gte=start_date)
            if end_date:
                requests_qs = requests_qs.filter(request_date__lte=end_date)

            ws.append([
                asset.asset_category,
                asset.model or "-",
                asset.serial_number or "-",
                requests_qs.count(),
                requests_qs.filter(status='approved').count()
            ])

    # ============================================================
    # 🔹 REPORT TYPE: REQUEST SUMMARY
    # ============================================================
    elif report_type == 'request_summary':

        ws.title = "Request Summary"

        headers = ["User", "Asset", "Model", "Request Date",
                   "Return Date", "Status", "Remarks"]
        ws.append(headers)

        for col, _ in enumerate(headers, 1):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).fill = header_fill

        requests = AssetRequest.objects.select_related('assigned_asset', 'user')

        if start_date:
            requests = requests.filter(request_date__gte=start_date)
        if end_date:
            requests = requests.filter(request_date__lte=end_date)
        if username:
            requests = requests.filter(user__username__icontains=username)

        for req in requests:
            asset = req.assigned_asset

            ws.append([
                req.user.username,
                asset.asset_category if asset else "-",
                asset.model if asset else "-",
                req.request_date.strftime("%Y-%m-%d"),
                req.return_date.strftime("%Y-%m-%d") if req.return_date else "-",
                req.status.capitalize(),
                req.remarks or "-"
            ])

    else:
        return HttpResponse("Invalid report type.", status=400)

    # -----------------------------
    # Auto-adjust column width
    # -----------------------------
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # -----------------------------
    # Filename Generator
    # -----------------------------
    filename = f"{report_type}_{timezone.localdate()}"
    if report_type == "request_summary" and username:
        filename += f"_{username}"
    filename += ".xlsx"

    # -----------------------------
    # Return Excel Response
    # -----------------------------
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response


















# sample data added
# Only allow admin/staff to generate
from django.contrib.auth.decorators import user_passes_test
def is_admin_or_staff(user):
    return user.is_authenticated and user.role in ['admin', 'staff']

@user_passes_test(is_admin_or_staff)
def generate_sample_assets(request):
    try:
        num = int(request.GET.get('num', 1000000))  
    except ValueError:
        num = 1000000

    for _ in range(num):
        AssetFactory()

    return HttpResponse(f"✅ Successfully created {num} sample assets!")




